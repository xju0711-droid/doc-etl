"""
复审模块主逻辑。

可插拔：完全独立于主提取流程，只读取 output/*.json。
可重复运行：每次复审刷新 results.xlsx 里的"复审报告" Sheet，不影响原始数据 Sheet。

支持：
  - 单文件 JSON（run 命令产出）
  - 列表 JSON（batch 命令产出的 batch_results.json）
  - 混合传入（自动按 source 字段去重，防止同一文档被统计两次）
"""

import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table

from .validators import validate

logger = logging.getLogger(__name__)
console = Console()

SHEET_NAME = "复审报告"

FILL_OK      = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_EMPTY   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_FMT     = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_SUMMARY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


# ── 数据加载 ──────────────────────────────────────────────────

def _load_records(paths: list[Path]) -> list[dict]:
    """
    读取 JSON 文件，返回去重后的记录列表。
    按 source 字段去重：同一份原始文档只保留最新提取结果。
    """
    seen: dict[str, dict] = {}  # source → record，后出现的覆盖先出现的

    for p in paths:
        try:
            raw = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning("无法读取 %s: %s", p, e)
            continue

        items = raw if isinstance(raw, list) else [raw]
        for item in items:
            if not isinstance(item, dict):
                continue
            if not item.get("data") or item.get("error"):
                continue  # 跳过失败记录
            src = item.get("source", str(p))
            seen[src] = item  # 相同 source 后者覆盖前者

    records = list(seen.values())
    logger.debug("加载 %d 份记录（来自 %d 个文件）", len(records), len(paths))
    return records


def _all_fields(records: list[dict]) -> list[str]:
    """取所有记录字段的并集（排除内部字段 _raw）。"""
    fields: set[str] = set()
    for r in records:
        fields.update(k for k in r["data"] if k != "_raw")
    return sorted(fields)


# ── 主入口 ────────────────────────────────────────────────────

def run_review(paths: list[Path], excel_path: Path) -> None:
    """
    对指定 JSON 文件执行复审，输出终端报告并写入 Excel。

    paths      : 要复审的 JSON 文件列表
    excel_path : 目标 Excel 文件（复审结果写入其中的"复审报告" Sheet）
    """
    records = _load_records(paths)
    if not records:
        console.print("[yellow]没有找到可复审的提取数据。请先运行 run 或 batch 命令。[/yellow]")
        return

    fields = _all_fields(records)
    logger.info("开始复审 %d 份文档，%d 个字段", len(records), len(fields))

    # 对每条记录的每个字段做校验
    rows = []
    for r in records:
        checks = {f: validate(f, r["data"].get(f)) for f in fields}
        rows.append({"source": r.get("source", "?"), "checks": checks})

    _print_terminal(rows, fields)
    _write_excel(rows, fields, excel_path)

    total_issues = sum(
        1 for row in rows
        for chk in row["checks"].values()
        if not chk.ok
    )

    if total_issues == 0:
        console.print("\n[green]所有字段均通过校验。[/green]")
    else:
        console.print(f"\n[yellow]共发现 {total_issues} 项问题（空值或格式异常）。[/yellow]")
    console.print(f"[dim]复审结果已写入: {excel_path}  →  '{SHEET_NAME}' Sheet[/dim]")


# ── 终端报告 ──────────────────────────────────────────────────

def _print_terminal(rows: list[dict], fields: list[str]) -> None:
    table = Table(
        title=f"复审报告  {len(rows)} 份文档 × {len(fields)} 个字段",
        show_lines=False,
    )
    table.add_column("来源文件", style="cyan", min_width=18, no_wrap=True)
    for f in fields:
        table.add_column(f, min_width=10)
    table.add_column("问题数", min_width=6)

    ok_counts = {f: 0 for f in fields}

    for row in rows:
        cells = [Path(row["source"]).name]
        n_issues = 0

        for f in fields:
            chk = row["checks"][f]
            if chk.ok:
                cells.append("[green]✓[/green]")
                ok_counts[f] += 1
            elif chk.status == "empty":
                cells.append("[red]✗ 空值[/red]")
                n_issues += 1
            else:
                cells.append(f"[yellow]⚠ {chk.note}[/yellow]")
                n_issues += 1

        issue_str = "[green]0[/green]" if n_issues == 0 else f"[red]{n_issues}[/red]"
        cells.append(issue_str)
        table.add_row(*cells)

    # 填充率汇总行
    fill_cells = ["[dim]字段填充率[/dim]"]
    for f in fields:
        rate = int(ok_counts[f] / len(rows) * 100) if rows else 0
        color = "green" if rate >= 80 else ("yellow" if rate >= 50 else "red")
        fill_cells.append(f"[{color}]{rate}%[/{color}]")
    fill_cells.append("")
    table.add_row(*fill_cells)

    console.print(table)
    console.print("[dim]图例: ✓ 正常  ✗ 空值  ⚠ 格式可疑[/dim]")


# ── Excel 报告 ────────────────────────────────────────────────

def _write_excel(rows: list[dict], fields: list[str], excel_path: Path) -> None:
    """刷新 Excel 中的"复审报告" Sheet（已有则替换，不影响其他 Sheet）。"""
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # 替换旧的复审报告 Sheet
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(title=SHEET_NAME, index=0)  # 放在最前面，方便查看

    all_cols = ["来源文件"] + fields + ["问题数", "复审时间"]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 表头
    for ci, name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B2"

    # 数据行
    for ri, row in enumerate(rows, 2):
        n_issues = 0

        ws.cell(row=ri, column=1, value=Path(row["source"]).name).alignment = Alignment(vertical="top")

        for fi, f in enumerate(fields, 2):
            chk = row["checks"][f]
            cell = ws.cell(row=ri, column=fi)
            cell.alignment = Alignment(horizontal="center", vertical="top")

            if chk.ok:
                cell.value = "✓"
                cell.fill = FILL_OK
            elif chk.status == "empty":
                cell.value = "✗ 空值"
                cell.fill = FILL_EMPTY
                n_issues += 1
            else:
                cell.value = f"⚠ {chk.note}"
                cell.fill = FILL_FMT
                n_issues += 1

        # 问题数列
        issue_col = len(fields) + 2
        ic = ws.cell(row=ri, column=issue_col, value=n_issues)
        ic.fill = FILL_OK if n_issues == 0 else FILL_EMPTY
        ic.alignment = Alignment(horizontal="center")

        # 复审时间列
        ws.cell(row=ri, column=issue_col + 1, value=ts).alignment = Alignment(horizontal="center")

    # 填充率汇总行
    sr = len(rows) + 2
    ws.cell(row=sr, column=1, value="字段填充率").fill = FILL_SUMMARY
    for fi, f in enumerate(fields, 2):
        ok = sum(1 for row in rows if row["checks"][f].ok)
        rate = f"{int(ok / len(rows) * 100)}%" if rows else "0%"
        cell = ws.cell(row=sr, column=fi, value=rate)
        cell.fill = FILL_SUMMARY
        cell.alignment = Alignment(horizontal="center")

    # 自动列宽
    for ci, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(ci)
        max_len = len(col_name)
        for ri in range(2, sr + 1):
            v = ws.cell(row=ri, column=ci).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[col_letter].width = max_len + 4

    excel_path.parent.mkdir(exist_ok=True)
    wb.save(excel_path)
    logger.info("复审报告已写入 %s", excel_path)