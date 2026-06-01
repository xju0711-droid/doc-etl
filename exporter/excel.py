"""
Excel 导出模块。

设计思路：
  output/results.xlsx 是持久化文件，多次运行结果不断追加。
  每种字段组合（Schema）对应一个 Sheet，同类字段的文档自动归到一起。
  例如：用 -f name=姓名 -f email=邮箱 提取了 20 个文档，
        它们全部出现在 "email+name" 这个 Sheet 里。

格式：
  - 第一行：深蓝色表头（元数据 + 字段名）
  - 后续行：每个文档一行，绿/黄/红背景表示质量分高低
  - 冻结首行 + 自动列宽
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_GOOD = "C6EFCE"   # 质量分 >= 80
COLOR_WARN = "FFEB9C"   # 质量分 60-79
COLOR_BAD  = "FFC7CE"   # 质量分 < 60 或失败

META_COLS = ["来源文件", "质量分", "是否通过", "使用模型"]


def _sheet_name(field_names: list[str]) -> str:
    """按字段名（排序后）生成 Sheet 名，Excel 上限 31 字符。"""
    name = "+".join(sorted(field_names))
    return name[:31]


def _apply_header(cell, text: str):
    cell.value = text
    cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    cell.font = Font(color=COLOR_HEADER_FG, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _row_fill(score) -> PatternFill:
    if score is None:
        color = COLOR_BAD
    elif score >= 80:
        color = COLOR_GOOD
    elif score >= 60:
        color = COLOR_WARN
    else:
        color = COLOR_BAD
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _write_headers(ws, field_names: list[str]):
    all_cols = META_COLS + field_names
    for col_idx, name in enumerate(all_cols, 1):
        _apply_header(ws.cell(row=1, column=col_idx), name)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _as_cell_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def append_to_excel(jobs: list, schema, output_path: Path) -> Path:
    """
    把一批任务结果追加写入 Excel 文件。

    - 文件不存在 → 新建
    - 对应 Sheet 不存在 → 新建 Sheet 并写表头
    - 对应 Sheet 已存在 → 直接在末尾追加数据行
    - 相同字段组合的文档始终归到同一个 Sheet
    """
    from queue_worker.worker import JobStatus

    field_names = list(schema.fields.keys())
    sheet_name = _sheet_name(field_names)

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认空 Sheet

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(title=sheet_name)
        _write_headers(ws, field_names)
        next_row = 2

    all_col_count = len(META_COLS) + len(field_names)

    for job in jobs:
        score = job.quality["score"] if job.quality else None
        fill = _row_fill(score)

        if job.status != JobStatus.DONE or job.result is None:
            row_values = [job.source, "失败", "否", "—"] + ["—"] * len(field_names)
        else:
            q = job.quality
            r = job.result
            row_values = [
                job.source,
                q["score"],
                "是" if q["passed"] else "否",
                r.model_used,
            ] + [_as_cell_value(r.data.get(fname)) for fname in field_names]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        next_row += 1

    # 自动调整列宽（扫描所有数据行，上限 55 字符避免太宽）
    all_cols = META_COLS + field_names
    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 55))
        ws.column_dimensions[col_letter].width = max_len + 4

    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    return output_path